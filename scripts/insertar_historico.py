#!/usr/bin/env python3
"""
Inserción del histórico 2025 en Supabase BirdLog usando supabase-py.
Lee el Excel y sube meteo y lindus por lotes directamente.

Uso:
    .venv/bin/python3 scripts/insertar_historico.py
"""
import sys
import openpyxl
from pathlib import Path
from datetime import date, time, datetime

import os
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")
SUPABASE_URL = os.environ["SUPABASE_DEV_URL"]
SUPABASE_ANON_KEY = os.environ["SUPABASE_DEV_KEY"]
EXCEL_PATH = Path(__file__).parent.parent / "docs" / "BirdLog_tablas_cliente_v03.xlsx"
BATCH_SIZE = 200

VIENTO_MAP = {
    "O": "W", "NO": "NW", "SO": "SW", "NNO": "NNW",
    "ONO": "WNW", "OSO": "WSW", "SSO": "SSW",
}
VIENTO_VALIDOS = {
    "N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
    "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW",
}


def normalizar_viento(v):
    if v is None:
        return None, False
    s = str(v).strip()
    if s in VIENTO_VALIDOS:
        return s, False
    if s in VIENTO_MAP:
        return VIENTO_MAP[s], False
    return None, True


def fmt_time(v):
    if v is None:
        return None
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime):
        return v.strftime("%H:%M:%S")
    return str(v)


def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def to_int(v):
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def load_id_maps(client):
    """Carga los mapas codigo_origen → id de visitas y especies."""
    visitas = client.table("visitas").select("id_visita,codigo_origen").execute()
    esp = client.table("especies").select("id_especie,codigo_origen").execute()
    v_map = {r["codigo_origen"]: r["id_visita"] for r in visitas.data}
    e_map = {r["codigo_origen"]: r["id_especie"] for r in esp.data}
    return v_map, e_map


def insertar_meteo(client, wb, v_map):
    ws = wb["meteo"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}

    total = 0
    lote = []
    lote_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_visita = row[h["id_visita"]]
        if cod_visita is None or str(cod_visita) not in v_map:
            continue  # V0001 y otras sin visita en BD

        id_visita = v_map[str(cod_visita)]
        cod = row[h["id_meteo"]]
        hora = fmt_time(row[h["hora"]])
        if hora is None:
            continue

        temp = to_float(row[h["temperatura"]])
        nub_raw = to_int(row[h["total_nubes_suma"]])
        nub = nub_raw if nub_raw is not None and 0 <= nub_raw <= 8 else None
        viento_raw = row[h["viento_suelo_direccion"]]
        viento_norm, es_comp = normalizar_viento(viento_raw)
        v_int = to_int(row[h["viento_suelo_fuerza"]])
        precip_raw = row[h["precipitaciones_cantidad"]]
        precip = str(precip_raw) if precip_raw is not None else None
        visib_raw = row[h["visibilidad"]]
        visib = to_int(visib_raw)
        presentes = to_int(row[h["presentes"]])
        observando = to_int(row[h["observando"]])
        visitantes = to_int(row[h["visitantes"]])
        obs = row[h["observaciones"]]
        if isinstance(obs, (int, float)):
            obs = str(obs)
        if es_comp:
            nota = f"viento: {viento_raw}"
            obs = f"{obs} | {nota}" if obs else nota
        hum = to_float(row[h["humedad_relativa"]])
        pres = to_float(row[h["presion_atm"]])
        precip_tipo = row[h["precipitaciones_tipo"]]
        if isinstance(precip_tipo, (int, float)):
            precip_tipo = str(precip_tipo)
        mn_cob = to_int(row[h["mar_nubes_cobertura"]])
        mn_alt = to_float(row[h["mar_nubes_altura"]])
        n1_cob = to_int(row[h["nubes_n1_cobertura"]])
        n1_alt = to_float(row[h["nubes_n1_altura"]])
        n1_tip = row[h["nubes_n1_tipo"]]
        n2_cob = to_int(row[h["nubes_n2_cobertura"]])
        n2_tip = row[h["nubes_n2_tipo"]]

        rec = {
            "id_visita": id_visita,
            "hora": hora,
            "temperatura": temp,
            "nubosidad": nub,
            "viento_direccion": viento_norm,
            "viento_intensidad": str(v_int) if v_int is not None else None,
            "precipitacion": precip,
            "visibilidad": str(visib) if visib is not None else None,
            "presentes": presentes,
            "observando": observando,
            "visitantes": visitantes,
            "observaciones": obs,
            "humedad_relativa": hum,
            "presion_atm": pres,
            "precipitacion_tipo": precip_tipo,
            "mar_nubes_cobertura": mn_cob,
            "mar_nubes_altura": mn_alt,
            "nubes_n1_cobertura": n1_cob,
            "nubes_n1_altura": n1_alt,
            "nubes_n1_tipo": n1_tip,
            "nubes_n2_cobertura": n2_cob,
            "nubes_n2_tipo": n2_tip,
            "codigo_origen": str(cod) if cod else None,
        }
        lote.append(rec)

        if len(lote) >= BATCH_SIZE:
            lote_num += 1
            print(f"  Meteo lote {lote_num} ({len(lote)} filas)...", end=" ", flush=True)
            res = client.table("meteorologia").insert(lote).execute()
            print(f"OK ({len(res.data)} insertadas)")
            total += len(res.data)
            lote = []

    if lote:
        lote_num += 1
        print(f"  Meteo lote {lote_num} ({len(lote)} filas)...", end=" ", flush=True)
        res = client.table("meteorologia").insert(lote).execute()
        print(f"OK ({len(res.data)} insertadas)")
        total += len(res.data)

    return total


def insertar_lindus(client, wb, v_map, e_map):
    ws = wb["Lindus_Trona"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}

    COMPORT = [("migrador", "MIGRADOR"), ("direccion_norte", "NORTE"), ("local", "LOCAL")]
    total = 0
    lote = []
    lote_num = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_visita = str(row[h["id_visita"]]) if row[h["id_visita"]] else None
        if not cod_visita or cod_visita == "V0001" or cod_visita not in v_map:
            continue
        cod_sp = str(row[h["id_especie"]]) if row[h["id_especie"]] else None
        if not cod_sp or cod_sp not in e_map:
            continue

        id_visita = v_map[cod_visita]
        id_especie = e_map[cod_sp]
        cod_obs = str(row[h["id_observacion"]])
        hora = fmt_time(row[h["hora_solar"]])
        if hora is None:
            continue
        edad = row[h["edad"]]
        sexo = row[h["sexo"]]
        plumaje = row[h["plumaje"]]
        obs = row[h["observaciones"]]
        sp_texto = row[h["especie_texto"]]

        partes = []
        for col, comport in COMPORT:
            n = to_int(row[h[col]]) or 0
            if n > 0:
                partes.append((comport, n))

        if not partes:
            continue

        for idx, (comport, numero) in enumerate(partes):
            cod_fila = f"{cod_obs}_{idx + 1}" if len(partes) > 1 else cod_obs
            lote.append({
                "id_visita": id_visita,
                "id_especie": id_especie,
                "hora": hora,
                "numero": numero,
                "comportamiento": comport,
                "edad": edad,
                "sexo": sexo,
                "plumaje": plumaje,
                "observaciones": obs,
                "especie_texto": sp_texto,
                "codigo_origen": cod_fila,
            })

            if len(lote) >= BATCH_SIZE:
                lote_num += 1
                print(f"  Lindus lote {lote_num} ({len(lote)} filas)...", end=" ", flush=True)
                res = client.table("lindus").insert(lote).execute()
                print(f"OK ({len(res.data)} insertadas)")
                total += len(res.data)
                lote = []

    if lote:
        lote_num += 1
        print(f"  Lindus lote {lote_num} ({len(lote)} filas)...", end=" ", flush=True)
        res = client.table("lindus").insert(lote).execute()
        print(f"OK ({len(res.data)} insertadas)")
        total += len(res.data)

    return total


def main():
    from supabase import create_client
    client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)

    print("Cargando mapas de IDs desde Supabase...")
    v_map, e_map = load_id_maps(client)
    print(f"  Visitas en BD: {len(v_map)}, Especies en BD: {len(e_map)}")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    print("\nInsertando METEOROLOGÍA...")
    total_meteo = insertar_meteo(client, wb, v_map)
    print(f"  Total meteo: {total_meteo} registros\n")

    print("Insertando LINDUS...")
    total_lindus = insertar_lindus(client, wb, v_map, e_map)
    print(f"  Total lindus: {total_lindus} registros\n")

    print("Verificación final:")
    for tabla in ["especies", "observadores", "lugares", "visitas", "meteorologia", "lindus"]:
        r = client.table(tabla).select("*", count="exact").limit(0).execute()
        print(f"  {tabla}: {r.count}")


if __name__ == "__main__":
    main()
