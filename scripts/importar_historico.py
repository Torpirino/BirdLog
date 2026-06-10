#!/usr/bin/env python3
"""
Importación del histórico 2025 desde BirdLog_tablas_cliente_v03.xlsx.
Genera SQL INSERT por tabla y lo imprime en bloques para ejecutar en Supabase.

Uso:
    .venv/bin/python3 scripts/importar_historico.py > /tmp/historico_import.sql

Decisiones aplicadas:
- utm_x/utm_y de Lindus/Trona → NULL (pendiente del cliente)
- id_observador → 'Gabi' para todas las visitas históricas
- tipo_visita → 'LINDUS' para las 98 visitas
- V0001 (sin id_lugar) → OMITIDA, sus 34 lindus también
- Fila mixta L002724 → dividida en 2 filas
- Viento compuesto o desconocido → NULL en viento_direccion, literal en observaciones
- Rumbos españoles (O, NO, SO, NNO, etc.) → normalizados a inglés
"""
import sys
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / "docs" / "BirdLog_tablas_cliente_v03.xlsx"

# Normalización de rumbos de viento español → 16 puntos en inglés
VIENTO_MAP = {
    "O": "W", "NO": "NW", "SO": "SW", "NNO": "NNW",
    "ONO": "WNW", "OSO": "WSW", "SSO": "SSW",
}
VIENTO_VALIDOS = {
    "N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
    "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW",
}


def esc(v):
    """Escapa un valor para SQL: NULL si None, texto con comillas simples."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    # date / time / datetime
    s = str(v)
    if hasattr(v, "strftime"):
        from datetime import date, time, datetime
        if isinstance(v, datetime):
            s = v.strftime("%Y-%m-%d")
        elif isinstance(v, date):
            s = v.strftime("%Y-%m-%d")
        elif isinstance(v, time):
            s = v.strftime("%H:%M:%S")
    # escape single quotes
    s = s.replace("'", "''")
    return f"'{s}'"


def normalizar_viento(v):
    """Devuelve (viento_normalizado, es_compuesto)."""
    if v is None:
        return None, False
    s = str(v).strip()
    if s in VIENTO_VALIDOS:
        return s, False
    if s in VIENTO_MAP:
        return VIENTO_MAP[s], False
    return None, True  # compuesto/irreconocible → NULL


def load_wb():
    return openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)


# ──────────────────────────────────────────────
# 1. ESPECIES
# ──────────────────────────────────────────────
def sql_especies(wb):
    lines = ["-- =================== ESPECIES ==================="]
    ws = wb["especies"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nc = row[h["nombre_cientifico"]]
        if nc is None:
            continue
        nc = str(nc).strip()
        nom = row[h["nombre_comun"]]
        grupo = row[h["grupo"]]
        cod = row[h["id_especie"]]
        rows.append(
            f"({esc(nc)}, {esc(nom)}, {esc(grupo)}, {esc(str(cod) if cod else None)})"
        )
    if rows:
        lines.append(
            "INSERT INTO especies (nombre_cientifico, nombre_comun, grupo, codigo_origen) VALUES"
        )
        lines.append(",\n".join(rows) + ";")
    return "\n".join(lines)


# ──────────────────────────────────────────────
# 2. OBSERVADORES
# ──────────────────────────────────────────────
def sql_observadores():
    return (
        "-- =================== OBSERVADORES ===================\n"
        "INSERT INTO observadores (nombre_observador) VALUES\n"
        "('Gabi'),\n"
        "('Ander')\n"
        "ON CONFLICT (nombre_observador) DO NOTHING;"
    )


# ──────────────────────────────────────────────
# 3. LUGARES
# ──────────────────────────────────────────────
def sql_lugares(wb):
    lines = ["-- =================== LUGARES ==================="]
    ws = wb["lugares"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = row[h["id_lugar"]]
        nombre = row[h["nombre_lugar"]]
        if nombre is None:
            continue
        rows.append(
            f"({esc(str(nombre).strip())}, 'CONTEO_MIGRATORIO', NULL, NULL, {esc(str(cod) if cod else None)})"
        )
    if rows:
        lines.append(
            "INSERT INTO lugares (nombre_lugar, tipo_lugar, utm_x, utm_y, codigo_origen) VALUES"
        )
        lines.append(",\n".join(rows) + ";")
    return "\n".join(lines)


# ──────────────────────────────────────────────
# 4. VISITAS (omite V0001)
# ──────────────────────────────────────────────
def sql_visitas(wb):
    lines = ["-- =================== VISITAS ==================="]
    ws = wb["visitas"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}
    rows = []
    omitidas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = row[h["id_visita"]]
        id_lugar_cod = row[h["id_lugar"]]
        if id_lugar_cod is None:
            omitidas.append(cod)
            continue
        fecha = row[h["fecha"]]
        hora_ini = row[h["hora_inicio"]]
        hora_fin = row[h["hora_fin"]]
        obs = row[h["observaciones_meteo"]]
        rows.append(
            f"((SELECT id_lugar FROM lugares WHERE codigo_origen={esc(str(id_lugar_cod))}), "
            f"(SELECT id_observador FROM observadores WHERE nombre_observador='Gabi'), "
            f"'LINDUS', {esc(fecha)}, {esc(hora_ini)}, {esc(hora_fin)}, {esc(obs)}, {esc(str(cod))})"
        )
    if rows:
        lines.append(
            "INSERT INTO visitas (id_lugar, id_observador, tipo_visita, fecha, hora_inicio, hora_fin, observaciones, codigo_origen) VALUES"
        )
        lines.append(",\n".join(rows) + ";")
    if omitidas:
        lines.append(f"-- Visitas omitidas (sin id_lugar): {omitidas}")
    return "\n".join(lines)


# ──────────────────────────────────────────────
# 5. METEOROLOGÍA
# ──────────────────────────────────────────────
def sql_meteo_batch(wb, batch_size=200):
    ws = wb["meteo"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}

    cols = (
        "id_visita, hora, temperatura, nubosidad, viento_direccion, "
        "viento_intensidad, precipitacion, visibilidad, presentes, observando, visitantes, "
        "observaciones, humedad_relativa, presion_atm, precipitacion_tipo, "
        "mar_nubes_cobertura, mar_nubes_altura, nubes_n1_cobertura, nubes_n1_altura, "
        "nubes_n1_tipo, nubes_n2_cobertura, nubes_n2_tipo, codigo_origen"
    )

    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_visita = row[h["id_visita"]]
        if cod_visita is None:
            continue
        cod = row[h["id_meteo"]]
        hora = row[h["hora"]]
        temp = row[h["temperatura"]]
        nub_raw = row[h["total_nubes_suma"]]
        nub = int(nub_raw) if nub_raw is not None else None
        if nub is not None and not (0 <= nub <= 8):
            nub = None
        viento_raw = row[h["viento_suelo_direccion"]]
        viento_norm, es_comp = normalizar_viento(viento_raw)
        v_int = row[h["viento_suelo_fuerza"]]
        precip = row[h["precipitaciones_cantidad"]]
        visib = row[h["visibilidad"]]
        presentes = row[h["presentes"]]
        observando = row[h["observando"]]
        visitantes = row[h["visitantes"]]
        obs = row[h["observaciones"]]
        if es_comp:
            nota = f"viento: {viento_raw}"
            obs = f"{obs} | {nota}" if obs else nota
        hum = row[h["humedad_relativa"]]
        pres = row[h["presion_atm"]]
        precip_tipo = row[h["precipitaciones_tipo"]]
        mn_cob = row[h["mar_nubes_cobertura"]]
        mn_alt = row[h["mar_nubes_altura"]]
        n1_cob = row[h["nubes_n1_cobertura"]]
        n1_alt = row[h["nubes_n1_altura"]]
        n1_tip = row[h["nubes_n1_tipo"]]
        n2_cob = row[h["nubes_n2_cobertura"]]
        n2_tip = row[h["nubes_n2_tipo"]]

        all_rows.append(
            f"((SELECT id_visita FROM visitas WHERE codigo_origen={esc(str(cod_visita))}), "
            f"{esc(hora)}, {esc(temp)}, {esc(nub)}, {esc(viento_norm)}, "
            f"{esc(v_int)}, {esc(precip)}, {esc(visib)}, "
            f"{esc(presentes)}, {esc(observando)}, {esc(visitantes)}, "
            f"{esc(obs)}, {esc(hum)}, {esc(pres)}, {esc(precip_tipo)}, "
            f"{esc(mn_cob)}, {esc(mn_alt)}, {esc(n1_cob)}, {esc(n1_alt)}, "
            f"{esc(n1_tip)}, {esc(n2_cob)}, {esc(n2_tip)}, {esc(str(cod) if cod else None)})"
        )

    batches = []
    for i in range(0, len(all_rows), batch_size):
        chunk = all_rows[i:i + batch_size]
        sql = f"INSERT INTO meteorologia ({cols}) VALUES\n" + ",\n".join(chunk) + ";"
        batches.append(sql)
    return batches


# ──────────────────────────────────────────────
# 6. LINDUS
# ──────────────────────────────────────────────
def sql_lindus_batch(wb, batch_size=500):
    ws = wb["Lindus_Trona"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {k: i for i, k in enumerate(headers)}

    cols = (
        "id_visita, id_especie, hora, numero, comportamiento, "
        "edad, sexo, plumaje, observaciones, especie_texto, codigo_origen"
    )

    COMPORT = [("migrador", "MIGRADOR"), ("direccion_norte", "NORTE"), ("local", "LOCAL")]
    all_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_visita = row[h["id_visita"]]
        # Omitir V0001 y cualquier fila sin visita
        if cod_visita is None or str(cod_visita) == "V0001":
            continue
        cod_obs = str(row[h["id_observacion"]])
        cod_sp = row[h["id_especie"]]
        hora = row[h["hora_solar"]]
        edad = row[h["edad"]]
        sexo = row[h["sexo"]]
        plumaje = row[h["plumaje"]]
        obs = row[h["observaciones"]]
        sp_texto = row[h["especie_texto"]]

        # Construir filas por comportamiento
        partes = []
        for col, comport in COMPORT:
            n = row[h[col]] or 0
            if n > 0:
                partes.append((comport, n))

        if not partes:
            continue  # no debería ocurrir

        for idx, (comport, numero) in enumerate(partes):
            # Si hay más de una parte → fila mixta, usar codigo_origen extendido
            if len(partes) > 1:
                cod_fila = f"{cod_obs}_{idx + 1}"
            else:
                cod_fila = cod_obs

            all_rows.append(
                f"((SELECT id_visita FROM visitas WHERE codigo_origen={esc(str(cod_visita))}), "
                f"(SELECT id_especie FROM especies WHERE codigo_origen={esc(str(cod_sp))}), "
                f"{esc(hora)}, {esc(numero)}, {esc(comport)}, "
                f"{esc(edad)}, {esc(sexo)}, {esc(plumaje)}, {esc(obs)}, "
                f"{esc(sp_texto)}, {esc(cod_fila)})"
            )

    batches = []
    for i in range(0, len(all_rows), batch_size):
        chunk = all_rows[i:i + batch_size]
        sql = f"INSERT INTO lindus ({cols}) VALUES\n" + ",\n".join(chunk) + ";"
        batches.append(sql)
    return batches


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main():
    wb = load_wb()

    print(sql_especies(wb))
    print()
    print(sql_observadores())
    print()
    print(sql_lugares(wb))
    print()
    print(sql_visitas(wb))
    print()

    meteo_batches = sql_meteo_batch(wb)
    print(f"-- =================== METEOROLOGÍA ({len(meteo_batches)} lotes) ===================")
    for i, batch in enumerate(meteo_batches, 1):
        print(f"-- Lote meteo {i}/{len(meteo_batches)}")
        print(batch)
        print()

    lindus_batches = sql_lindus_batch(wb)
    print(f"-- =================== LINDUS ({len(lindus_batches)} lotes) ===================")
    for i, batch in enumerate(lindus_batches, 1):
        print(f"-- Lote lindus {i}/{len(lindus_batches)}")
        print(batch)
        print()

    print("-- =================== VERIFICACIÓN ===================")
    print("""SELECT 'especies' AS tabla, COUNT(*) AS filas FROM especies
UNION ALL SELECT 'observadores', COUNT(*) FROM observadores
UNION ALL SELECT 'lugares', COUNT(*) FROM lugares
UNION ALL SELECT 'visitas', COUNT(*) FROM visitas
UNION ALL SELECT 'meteorologia', COUNT(*) FROM meteorologia
UNION ALL SELECT 'lindus', COUNT(*) FROM lindus;""")


if __name__ == "__main__":
    main()
